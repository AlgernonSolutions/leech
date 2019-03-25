import logging

from aws_xray_sdk.core import xray_recorder

from toll_booth.alg_tasks.rivers.rocks import task


@xray_recorder.capture('get_payroll_data')
@task('get_payroll_data')
def get_payroll_data(**kwargs):
    from toll_booth.alg_obj.aws.snakes.stored_statics import StaticCsv
    from toll_booth.alg_obj.forge.extractors.credible_fe import CredibleFrontEndDriver
    import dateutil

    credible_date_format = '%m/%d/%Y'
    id_source = kwargs['id_source']
    check_date = kwargs['check_date']
    check_date_value = dateutil.parser.parse(check_date)
    check_dates = StaticCsv.for_check_dates(id_source)
    check_row = check_dates[check_date_value][0]
    sample_start_date = check_row['Sample Start']
    sample_end_date = check_row['Sample End']
    check_row_index = check_dates.index(check_row)
    previous_row = check_dates.get_by_index(check_row_index - 1)
    previous_sample_date = previous_row['Sample Date']
    staff_search_data = {
        'emp_status_f': 'ACTIVE',
        'first_name': 1,
        'last_name': 1,
        'emp_id': 1,
        'profile_code': 1,
        'asgn_supervisors': 1,
        'asgn_supervisees': 1
    }
    client_search_data = {
        'teams': 1,
        'client_id': 1,
        'last_name': 1,
        'first_name': 1,
        'text28': 1,
        'dob': 1,
        'ssn': 1,
        'primary_assigned': 1,
        'client_status_f': 'ALL ACTIVE'
    }
    sample_search_data = {
        'clientvisit_id': 1,
        'service_type': 1,
        'consumer_name': 1,
        'staff_name': 1,
        'client_int_id': 1,
        'emp_int_id': 1,
        'non_billable1': 0,
        'visittype': 1,
        'orig_rate_amount': 1,
        'timein': 1,
        'timeout': 1,
        'orig_units': 1,
        'wh_fld1': 'cv.transfer_date',
        'wh_cmp1': '>=',
        'wh_val1': sample_start_date.strftime(credible_date_format),
        'wh_andor': 'AND',
        'wh_fld2': 'cv.transfer_date',
        'wh_cmp2': '<=',
        'wh_val2': sample_end_date.strftime(credible_date_format),
        'data_dict_ids': [80, 81, 82, 87, 83],
        'show_unappr': 1
    }
    recovery_search_data = sample_search_data.copy()
    recovery_search_data.update({
        'wh_fld1': 'cv.transfer_data',
        'wh_cmp1': '<',
        'wh_val1': sample_start_date.strftime(credible_date_format),
        'wh_fld2': 'cv.appr_date',
        'wh_cmp2': '>',
        'wh_val2': previous_sample_date.strftime(credible_date_format),
        'show_unappr': 0
    })
    report_args = {
        'emp_data': ('Employees', staff_search_data),
        'client_data': ('Clients', client_search_data),
        'sample_data': ('ClientVisit', sample_search_data),
        'recovery_data': ('ClientVisit', recovery_search_data)
    }
    results = {'sampled_days': (sample_end_date-sample_start_date).days}
    with CredibleFrontEndDriver(id_source) as driver:
        for report_name, report_arg in report_args.items():
            try:
                results[report_name] = driver.process_advanced_search(*report_arg)
            except RuntimeError:
                results[report_name] = []
    if not results:
        return
    return results


@xray_recorder.capture('get_productivity_report_data')
@task('get_productivity_report_data')
def get_productivity_report_data(**kwargs):
    from toll_booth.alg_obj.forge.extractors.credible_fe import CredibleFrontEndDriver
    from datetime import datetime, timedelta

    logging.info(f'called a get_productivity report data, kwargs: {kwargs}')
    credible_date_format = '%m/%d/%Y'

    today = datetime.utcnow()
    encounter_start_date = today - timedelta(days=90)
    unapproved_start_date = today - timedelta(days=365)
    staff_search_data = {
        'emp_status_f': 'ACTIVE',
        'first_name': 1,
        'last_name': 1,
        'emp_id': 1,
        'profile_code': 1,
        'asgn_supervisors': 1,
        'asgn_supervisees': 1
    }
    client_search_data = {
        'teams': 1,
        'client_id': 1,
        'last_name': 1,
        'first_name': 1,
        'text28': 1,
        'dob': 1,
        'ssn': 1,
        'primary_assigned': 1,
        'client_status_f': 'ALL ACTIVE'
    }
    encounter_search_data = {
        'clientvisit_id': 1,
        'service_type': 1,
        'non_billable': 1,
        'consumer_name': 1,
        'staff_name': 1,
        'client_int_id': 1,
        'emp_int_id': 1,
        'non_billable1': 3,
        'visittype': 1,
        'orig_rate_amount': 1,
        'timein': 1,
        'wh_fld1': 'cv.transfer_date',
        'wh_cmp1': '>=',
        'wh_val1': encounter_start_date.strftime(credible_date_format),
        'data_dict_ids': 83
    }
    unapproved_search_data = encounter_search_data.copy()
    unapproved_search_data.update({
        'non_billable1': 0,
        'show_unappr': 1,
        'wh_fld1': 'cv.appr',
        'wh_cmp1': '=',
        'wh_val1': False,
        'data_dict_ids': 641
    })
    tx_plan_args = encounter_search_data.copy()
    tx_plan_args['visittype_id'] = 3
    da_args = encounter_search_data.copy()
    da_args['visittype_id'] = 5

    id_source = kwargs['id_source']
    report_args = {
        'emp_data': ('Employees', staff_search_data),
        'client_data': ('Clients', client_search_data),
        'encounter_data': ('ClientVisit', encounter_search_data),
        'unapproved_data': ('ClientVisit', unapproved_search_data, unapproved_start_date, today),
        'tx_data': ('ClientVisit', tx_plan_args, unapproved_start_date, today),
        'da_data': ('ClientVisit', da_args, unapproved_start_date, today)
    }
    results = {}
    with CredibleFrontEndDriver(id_source) as driver:
        for report_name, report_arg in report_args.items():
            results[report_name] = driver.process_advanced_search(*report_arg)
    if not results:
        return
    return results


@xray_recorder.capture('build_daily_report')
@task('build_daily_report')
def build_daily_report(**kwargs):
    import re
    from decimal import Decimal

    daily_report = {}
    encounter_data = kwargs['encounter_data']
    encounters = [{
        'clientvisit_id': int(x['Service ID']),
        'transfer_date': x['Transfer Date'],
        'visit_type': x['Service Type'],
        'non_billable': x['Non Billable'] == 'True',
        'emp_id': int(x['Staff ID']),
        'client_id': int(x['Consumer ID']),
        'base_rate': Decimal(re.sub(r'[^\d.]', '', x['Base Rate'])),
        'data_dict_ids': 83
    } for x in encounter_data]
    unapproved_data = kwargs['unapproved_data']
    unapproved = [{
        'clientvisit_id': int(x['Service ID']),
        'rev_timeout': x['Service Date'],
        'visit_type': x['Service Type'],
        'non_billable': bool(x['Non Billable']),
        'emp_id': int(x['Staff ID']),
        'client_id': int(x['Consumer ID']),
        'red_x': x['Manual RedX Note'],
        'base_rate': Decimal(re.sub(r'[^\d.]', '', x['Base Rate']))
    } for x in unapproved_data]
    tx_plan_data = kwargs['tx_data']
    tx_plans = [{
        'rev_timeout': x['Service Date'],
        'emp_id':  int(x['Staff ID']),
        'client_id': int(x['Consumer ID'])
    } for x in tx_plan_data]
    da_data = kwargs['da_data']
    diagnostics = [{
        'rev_timeout': x['Service Date'],
        'emp_id': int(x['Staff ID']),
        'client_id': int(x['Consumer ID'])
    } for x in da_data]
    caseloads = kwargs['caseloads']
    for team_name, employees in caseloads.items():
        if team_name == 'unassigned':
            continue
        page_name = f'productivity_{team_name}'
        productivity_results = _build_team_productivity(employees, encounters, unapproved)
        daily_report[page_name] = productivity_results
    tx_report = _build_expiration_report(caseloads, tx_plans, 180)
    da_report = _build_expiration_report(caseloads, diagnostics, 180)
    thirty_sixty_ninety = _build_not_seen_report(caseloads, encounters)
    unassigned_report = _build_unassigned_report(caseloads)
    daily_report['tx_plans'] = tx_report
    daily_report['diagnostics'] = da_report
    daily_report['unassigned'] = unassigned_report
    daily_report['30, 60, 90'] = thirty_sixty_ninety
    return {'report_data': daily_report}


@xray_recorder.capture('get_da_tx_data')
@task('get_da_tx_data')
def get_da_tx_data(**kwargs):
    from toll_booth.alg_obj.forge.extractors.credible_fe import CredibleFrontEndDriver

    id_source = kwargs['id_source']
    with CredibleFrontEndDriver(id_source) as driver:
        pass


@xray_recorder.capture('write_report_data')
@task('write_report_data')
def write_report_data(**kwargs):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    from toll_booth.alg_obj.aws.snakes.invites import ObjectDownloadLink
    import os
    from datetime import datetime

    report_bucket_name = kwargs.get('report_bucket_name', os.getenv('REPORT_BUCKET_NAME', 'algernonsolutions-leech'))
    local_user_directory = os.path.expanduser('~')
    report_name = kwargs['report_name']
    id_source = kwargs['id_source']
    today = datetime.utcnow()
    today_string = today.strftime('%Y%m%d')
    report_name = f'{report_name}_{today_string}.xlsx'
    report_save_path = os.path.join(local_user_directory, report_name)
    reports = kwargs['report_data']
    report_book = Workbook()
    front_sheet = report_book.active
    for entry_name, report_data in reports.items():
        new_sheet = report_book.create_sheet(entry_name)
        new_sheet.append([entry_name])
        top_row = new_sheet.row_dimensions[1]
        top_row.font = Font(bold=True, size=18)
        new_sheet.append([])
        row_lengths = [len(x) for x in report_data]
        max_row_length = None
        if row_lengths:
            max_row_length = max([len(x) for x in report_data])
        for row in report_data:
            new_sheet.append(row)
        if max_row_length:
            new_sheet.merge_cells(f'A1:{chr(ord("a") + (max_row_length-1))}1')
    report_book.remove(front_sheet)
    try:
        report_book.save(report_save_path)
    except FileNotFoundError:
        report_save_path = f'/tmp/{report_name}'
        report_book.save(report_save_path)
    download_link = ObjectDownloadLink(report_bucket_name, f'{id_source}/{report_name}', local_file_path=report_save_path)
    return {'download_link': download_link}


@xray_recorder.capture('send_report')
@task('send_report')
def send_report(**kwargs):
    from toll_booth.alg_obj.aws.snakes.stored_statics import StaticJson

    report_recipients = StaticJson.for_report_recipients(**kwargs).stored_asset
    download_link = kwargs['download_link']
    subject_line = 'Algernon Solutions Clinical Intelligence Report'
    text_body = f''' 
        You are receiving this email because you have requested to have routine reports sent to you through the Algernon Clinical Intelligence Platform. 
        The requested report can be downloaded from the included link. To secure the information contained within, the link will expire in {download_link.expiration_hours} hours.
        I hope this report brings you joy and the everlasting delights of a cold data set.

        {str(download_link)}

        - Algernon

        This communication, download link, and any attachment may contain information, which is sensitive, confidential and/or privileged, covered under HIPAA and is intended for use only by the addressee(s) indicated above. If you are not the intended recipient, please be advised that any disclosure, copying, distribution, or use of the contents of this information is strictly prohibited. If you have received this communication in error, please notify the sender immediately and destroy all copies of the original transmission.
    '''
    html_body = f'''
       <!DOCTYPE html>
        <html lang="en" xmlns="http://www.w3.org/1999/html" xmlns="http://www.w3.org/1999/html">
            <head>
                <meta charset="UTF-8">
                <title>Algernon Clinical Intelligence Report</title>
                <style>
                   
                    .container {{
                        position: relative;
                        text-align: center;
                    }}
                </style>
            </head>
            <body>
                <p>You are receiving this email because you have requested to have routine reports sent to you through the Algernon Clinical Intelligence Platform.</p>
                <p>The requested report can be downloaded from the included link. To secure the information contained within, the link will expire in {download_link.expiration_hours} hours.</p>
                <p>I hope this report brings you joy and the everlasting delights of a cold data set.</p>
                <h4><a href="{str(download_link)}">Download Report</a></h4>
                <p> - Algernon </p>
                <h5>
                    This communication, download link, and any attachment may contain information, which is sensitive, confidential and/or privileged, covered under HIPAA and is intended for use only by the addressee(s) indicated above.<br/>
                    If you are not the intended recipient, please be advised that any disclosure, copying, distribution, or use of the contents of this information is strictly prohibited.<br/>
                    If you have received this communication in error, please notify the sender immediately and destroy all copies of the original transmission.<br/>
                </h5>
            </body>
        </html>
    '''

    response = _send_by_ses(recipients=report_recipients['recipients'], subject_line=subject_line, text_body=text_body, html_body=html_body, **kwargs)
    return {'message_id': response['MessageId'], 'text_body': text_body, 'html_body': html_body}


@xray_recorder.capture('build_clinical_teams')
@task('build_clinical_teams')
def build_clinical_teams(**kwargs):
    from toll_booth.alg_obj.aws.snakes.stored_statics import StaticJson

    id_source = kwargs['id_source']
    team_json = StaticJson.for_team_data(id_source)
    teams = team_json['teams']
    manual_assignments = team_json['manual_assignments']
    first_level = team_json['first_level']
    default_team = team_json['default_team']
    emp_data = kwargs['emp_data']

    for entry in emp_data:
        int_emp_id = int(entry['Employee ID'])
        str_emp_id = str(int_emp_id)
        supervisor_names = entry['Supervisors']
        profile_code = entry['profile_code']
        if supervisor_names is None or profile_code != 'CSA Community Support Licensed NonLicensed':
            continue
        emp_record = {
            'emp_id': int_emp_id,
            'first_name': entry['First Name'],
            'last_name': entry['Last Name'],
            'profile_code': entry['profile_code'],
            'caseload': []
        }
        if str_emp_id in manual_assignments:
            teams[manual_assignments[str_emp_id]].append(emp_record)
            continue
        for name in first_level:
            if name in supervisor_names:
                teams[name].append(emp_record)
                break
        else:
            teams[default_team].append(emp_record)
    return {'teams': teams}


@xray_recorder.capture('build_clinical_caseloads')
@task('build_clinical_caseloads')
def build_clinical_caseloads(**kwargs):
    caseloads = {}
    unassigned = []
    name_lookup = {}
    teams = kwargs['teams']
    clients = kwargs['client_data']
    for team_name, employees in teams.items():
        if team_name not in caseloads:
            caseloads[team_name] = {}
        for employee in employees:
            emp_id = str(employee['emp_id'])
            last_name = employee['last_name']
            first_name = employee['first_name']
            list_name = f'{first_name[0]} {last_name}'
            name_lookup[list_name] = emp_id
            if emp_id not in caseloads[team_name]:
                caseloads[team_name][emp_id] = employee
    for client in clients:
        client_id = client[' Id']
        primary_assigned = client['Primary Staff']
        if not primary_assigned:
            unassigned.append({
                'client_id': client_id,
                'last_name': client['Last Name'],
                'first_name': client['First Name'],
                'medicaid_id': client['Medicaid ID'],
                'dob': client['DOB'],
                'ssn': client['SSN'],
                'team': client['CSA (Team)']
            })
            continue
        primary_names = _parse_staff_names(primary_assigned)
        client_record = {
            'client_id': client_id,
            'last_name': client['Last Name'],
            'first_name': client['First Name'],
            'medicaid_id': client['Medicaid ID'],
            'dob': client['DOB'],
            'ssn': client['SSN'],
            'primary_staff': primary_names
        }
        found = False
        for staff_name in primary_names:
            found_emp_id = name_lookup.get(staff_name)
            if found_emp_id:
                for team_name, employees in caseloads.items():
                    for emp_id, employee in employees.items():
                        if emp_id == found_emp_id:
                            employee['caseload'].append(client_record)
                            found = True
        if not found:
            unassigned.append({
                'client_id': client_id,
                'last_name': client['Last Name'],
                'first_name': client['First Name'],
                'medicaid_id': client['Medicaid ID'],
                'dob': client['DOB'],
                'ssn': client['SSN'],
                'team': client['CSA (Team)'],
                'primary_staff': primary_assigned
            })
    client_ids = set()
    for team_name, team in caseloads.items():
        for emp_id, employee in team.items():
            client_ids.update([x['client_id'] for x in employee['caseload']])
    client_ids.update([x['client_id'] for x in unassigned])
    if client_ids - set([str(x[' Id']) for x in clients]):
        raise RuntimeError('while creating caseloads, we seemed to have missed someone, can not continue due to prime directive')
    caseloads['unassigned'] = unassigned
    return {'caseloads': caseloads}


@xray_recorder.capture('build_payroll_report')
@task('build_payroll_report')
def build_payroll_report(**kwargs):
    from toll_booth.alg_obj.aws.snakes.stored_statics import StaticCsv

    header_line = [
        'Team Name', 'CSW Name', 'Service Date', 'Client ID', 'Service ID', 'Service Type',
        'Units', 'Transfer Date', 'Approved', 'Approved By', 'Approved Date',
        'Delay', 'Timing', 'Rate', 'Rate Payable'
    ]
    pay_csv = StaticCsv.for_pay_rates(kwargs['id_source'])
    payroll_report = {}
    sample_period_data = _standardize_encounter_data(kwargs['sample_data'])
    recovery_data = _standardize_encounter_data(kwargs['recovery_data'])
    for team_name, team in kwargs['teams'].items():
        team_report = [header_line]
        team_unapproved_report = [header_line]
        team_recovery_report = [header_line]
        for employee in team:
            employee_name = f'{employee["last_name"]}, {employee["first_name"]}'
            emp_id = employee['emp_id']
            try:
                pay_row = pay_csv[str(emp_id)][0]
            except KeyError:
                pay_row = {'Rate': '$0'}
            format_args = (team_name, emp_id, employee_name, pay_row, kwargs.get('sampled_days', 16))
            employee_encounters, employee_unapproved = _format_sampled_encounters(sample_period_data, *format_args, recovery=False)
            employee_recovery = _format_sampled_encounters(recovery_data, *format_args, recovery=True)
            team_report.extend(employee_encounters)
            team_unapproved_report.extend(employee_unapproved)
            team_recovery_report.extend(employee_recovery)
        payroll_report[f'payable_{team_name}'] = team_report
        payroll_report[f'unapproved_{team_name}'] = team_unapproved_report
        payroll_report[f'recovery_{team_name}'] = team_recovery_report
    return {'report_data': payroll_report}


def _format_sampled_encounters(sample_period_data, team_name, emp_id, employee_name, pay_row, sampled_days, recovery):
    from decimal import Decimal
    import re

    def match_appr_emp_id(test_row):
        return str(test_row['emp_id']) == str(emp_id) and test_row['appr'] is True

    def match_emp_id(test_row):
        return str(test_row['emp_id']) == str(emp_id)

    sampled_encounters = []
    unapproved_encounters = []
    average_daily_productivity = sum([int(x['base_units'])/4 for x in filter(match_appr_emp_id, sample_period_data)])/sampled_days
    base_rate = Decimal(re.sub(r'[^\d.]', '', pay_row['Rate']))
    if average_daily_productivity > 3.3 and not recovery:
        base_rate = 25.00
    for row in filter(match_emp_id, sample_period_data):
        service_rate = base_rate
        approved = row['appr']
        if not approved:
            service_rate = 0
        transfer_date = row['transfer_date']
        service_date = row['rev_timeout']
        delay_delta = (transfer_date-service_date)
        delay = ((delay_delta.days * 24) + (delay_delta.seconds//3600))
        base_units = row['base_units']
        timing = 'Late'
        if delay <= 48:
            timing = 'On Time'
            if delay <= 24:
                timing = 'Early'
                if service_rate:
                    service_rate += 2
        rate_payable = int(service_rate) * int(base_units) / 4
        claim_row = [
            team_name, employee_name, service_date, row['client_id'], row['clientvisit_id'],
            row['service_type'], int(base_units), transfer_date, approved, row['appr_user'], row['appr_date'],
            delay, timing, service_rate, rate_payable
        ]
        if approved:
            sampled_encounters.append(claim_row)
        else:
            unapproved_encounters.append(claim_row)
    if recovery:
        return sampled_encounters
    return sampled_encounters, unapproved_encounters


def _standardize_encounter_data(encounter_data):
    return [{
        'clientvisit_id': x['Service ID'],
        'rev_timeout': x['Time Out'],
        'service_date': x['Service Date'],
        'service_type': x['Service Type'],
        'emp_id': x['Staff ID'],
        'client_id': x['Consumer ID'],
        'appr': x['Approved'] == 'True',
        'appr_date': x['Approved Date'],
        'appr_user': x['Approved User'],
        'transfer_date': x['Transfer Date'],
        'base_units': x['Base Units']
    } for x in encounter_data]


def _parse_staff_names(primary_staff_line):
    import re

    staff = []
    if not primary_staff_line:
        return staff
    program_pattern = '(\([^)]*\))'
    program_re = re.compile(program_pattern)
    staff_names = primary_staff_line.split(', ')
    for name in staff_names:
        program_match = program_re.search(name)
        if program_match:
            program_name = program_match.group(1)
            name = name.replace(program_name, '')
        staff.append(name)
    return staff


def _build_team_productivity(team_caseload, encounters, unapproved):
    from datetime import datetime, timedelta

    results = [['CSW ID', 'CSW Name', 'Past 24 Hours', 'Next Past 6 Days', 'All Unapproved', 'All Red X Notes']]
    twenty_four_hours_ago = datetime.now() - timedelta(hours=24)
    six_days_ago = datetime.now() - timedelta(days=6)
    past_day_encounters = [x for x in encounters if x['transfer_date'] >= twenty_four_hours_ago]
    next_six_days_encounters = [x for x in encounters if all(
        [x['transfer_date'] < twenty_four_hours_ago, x['transfer_date'] >= six_days_ago]
    )]
    for emp_id, employee in team_caseload.items():
        emp_id = int(emp_id)
        emp_past_day_encounters = [x['base_rate'] for x in past_day_encounters if x['emp_id'] == emp_id]
        emp_next_six_days_encounters = [x['base_rate'] for x in next_six_days_encounters if x['emp_id'] == emp_id]
        emp_red_x = [x['base_rate'] for x in unapproved if x['emp_id'] == emp_id and x['red_x']]
        emp_unapproved = [x['base_rate'] for x in unapproved if x['emp_id'] == emp_id and not x['red_x']]
        results.append([
            emp_id, f'{employee["last_name"]}, {employee["first_name"]}',
            sum(emp_past_day_encounters), sum(emp_next_six_days_encounters), sum(emp_unapproved), sum(emp_red_x)
        ])
    return results


def _build_expiration_report(caseloads, assessment_data, assessment_lifespan):
    from datetime import datetime, timedelta

    lifespan_delta = timedelta(days=assessment_lifespan)
    inverted = _invert_caseloads(caseloads)
    now = datetime.now()
    max_assessments = {}
    results = [['Team', 'CSW Name', 'Client ID', 'Start Date', 'End Date', 'Expired', 'Days Remaining']]
    for assessment in assessment_data:
        client_id = str(assessment['client_id'])
        if client_id not in max_assessments:
            max_assessments[client_id] = []
        max_assessments[client_id].append(assessment['rev_timeout'])
    for client_id, assessments in max_assessments.items():
        assignments = inverted.get(client_id, {'team': 'unassigned', 'csw': 'unassigned'})
        team_name, csw_name = assignments['team'], assignments['csw']
        max_assessment_date = max(assessments)
        expiration_date = max_assessment_date + lifespan_delta
        expired = False
        days_left = (expiration_date - now).days
        if expiration_date < now:
            expired = True
            days_left = 0
        results.append([team_name, csw_name, client_id, max_assessment_date, expiration_date, expired, days_left])
    no_assessments = set(inverted.keys()) - set(max_assessments.keys())
    for client_id in no_assessments:
        assignments = inverted.get(client_id, {'team': 'unassigned', 'csw': 'unassigned'})
        team_name, csw_name = assignments['team'], assignments['csw']
        results.append([team_name, csw_name, client_id, '',  '', True, 0])
    return results


def _build_not_seen_report(caseloads, encounter_data):
    from datetime import datetime
    today = datetime.now()
    results = [['Team', 'CSW Name', 'Client ID', 'Last Service by CSW', 'Last Billable Service', '30/60/90 per CSW', '30/60/90 per Last Billable']]
    inverted = _invert_caseloads(caseloads)
    for client_id, assignments in inverted.items():
        team = assignments['team']
        if team == 'unassigned':
            continue
        csw_id = assignments['emp_id']
        csw_name = assignments['csw']
        client_encounters = [x for x in encounter_data if int(x['client_id']) == int(client_id)]
        if not client_encounters:
            results.append([team, csw_name, client_id, '?', '?', '90', '90'])
            continue
        max_encounter_date = max([x['rev_timeout'] for x in client_encounters])
        per_billable = _calculate_thirty_sixty_ninety(today, max_encounter_date)
        csw_encounters = [x for x in client_encounters if int(x['emp_id']) == int(csw_id)]
        if not csw_encounters:
            results.append([team, csw_name, client_id, '?', max_encounter_date, '90', per_billable])
            continue
        max_csw_date = max([x['rev_timeout'] for x in csw_encounters])
        per_csw = _calculate_thirty_sixty_ninety(today, max_csw_date)
        results.append([team, csw_name, client_id, max_csw_date, max_encounter_date, per_csw, per_billable])
    return results


def _calculate_thirty_sixty_ninety(today, max_encounter_date):
    encounter_age = (today - max_encounter_date).days
    if encounter_age <= 30:
        return '30'
    if encounter_age <= 60:
        return '60'
    return '90'


def _build_unassigned_report(caseloads):
    report = [['Client ID', 'Client Name', 'DOB', 'SSN', 'Medicaid Number', 'Assigned CSA', 'Primary Assigned Staff']]
    for client in caseloads['unassigned']:
        primary_staff = client.get('primary_staff')
        if primary_staff:
            if isinstance(primary_staff, list):
                primary_staff = ', '.join(primary_staff)
        report.append([
            client['client_id'], f'{client["last_name"]}, {client["first_name"]}',
            client['dob'], client['dob'], client['ssn'], client['team'], primary_staff
        ])
    return report


def _send_by_ses(**kwargs):
    import boto3

    client = boto3.client('ses')
    recipients = kwargs['recipients']
    html_body, text_body = kwargs['html_body'], kwargs['text_body']
    subject_line = kwargs['subject_line']

    response = client.send_email(
        Source='algernon@algernon.solutions',
        Destination={
            'ToAddresses': [x['email_address'] for x in recipients]
        },
        Message={
            'Subject': {'Data': subject_line},
            'Body': {
                'Text': {'Data': text_body},
                'Html': {'Data': html_body}
            }
        },
        ReplyToAddresses=['algernon@algernon.solutions']
    )
    return response


def _send_by_pinpoint(**kwargs):
    import os
    import boto3

    client = boto3.client('pinpoint')
    application_id = kwargs.get('application_id', os.getenv('PINPOINT_APP_ID', 'e077e1ece06a40d983a1fb0cdeb76854'))
    recipients = kwargs['recipients']
    html_body, text_body = kwargs['html_body'], kwargs['text_body']
    subject_line = kwargs['subject_line']
    addresses = {
        x['email_address']: {'ChannelType': 'EMAIL'} for x in recipients
    }

    response = client.send_messages(
        ApplicationId=application_id,
        MessageRequest={
            'Addresses': addresses,
            'MessageConfiguration': {
                'EmailMessage': {
                    'FromAddress': 'algernon@algernon.solutions',
                    'ReplyToAddresses': [
                        'algernon@algernon.solutions',
                    ],
                    'SimpleEmail': {
                        'HtmlPart': {
                            'Data': html_body
                        },
                        'Subject': {
                            'Data': subject_line
                        },
                        'TextPart': {
                            'Data': text_body
                        }
                    }
                }
            }
        }
    )
    return response


def _invert_caseloads(caseloads):
    inverted_caseloads = {}
    for team_name, team_caseload in caseloads.items():
        if team_name == 'unassigned':
            for client in team_caseload:
                inverted_caseloads[client['client_id']] = {'team': 'unassigned', 'csw': 'unassigned', 'emp_id': 0}
            continue
        for emp_id, employee in team_caseload.items():
            csw = f'{employee["last_name"]}, {employee["first_name"]}'
            for client in employee['caseload']:
                inverted_caseloads[client['client_id']] = {'team': team_name, 'csw': csw, 'emp_id': emp_id}
    return inverted_caseloads
